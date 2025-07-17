import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import PyPDF2
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import re
from datetime import datetime, time
import os
#PDFの内容から各種データを抽出し、データを基に氏名や出退勤日、
#時間をエクセルに自動で書き写すアプリ

#データ取得のPDFの選択と表示
def select_pdf():
    global pdf_filename
    pdf_filename = filedialog.askopenfilename(initialdir="./", title="Select a PDF File")
    pdf_label.config(text=f"選択されたPDF: {pdf_filename}")

#処理したいエクセルファイル格納フォルダ設定と表示
def select_excel_folder():
    global excel_folder
    excel_folder = filedialog.askdirectory(initialdir="./", title="Select Excel Folder")
    excel_label.config(text=f"選択されたフォルダ: {excel_folder}")

#実行しますか→はいを押下で処理スタート
def process_files():
    if not pdf_filename or not excel_folder:
        messagebox.showerror("エラー", "PDFファイルとエクセルフォルダを選択してください。")
        return

    if not messagebox.askyesno("確認", "実行しますか？"):
        return

    error_messages = []

    try:
        # PDFファイルを読み込む
        pdf_file = open(pdf_filename, 'rb')
        reader = PyPDF2.PdfReader(pdf_file)

        # ページ数を取得する
        page_count = len(reader.pages)

        # 進捗バーの設定
        progress_bar['maximum'] = page_count
        progress_bar['value'] = 0
        progress_label.config(text="処理中です...")

        # ページ毎にデータを取得する
        for page in range(page_count):

            # ページを取得する
            current_page = reader.pages[page]
            
            # ページ内のテキストデータを取得する
            content = current_page.extract_text()
            
            # 支援+10桁直後の2単語を抽出（氏名抽出）
            shien_pattern = re.compile(r'支援\d{10}\s+(\S+\s+\S+)')
            shien_matches = shien_pattern.findall(content)

            #氏名の後に数字や空白があれば削除処理
            shien_combined = ''.join(shien_matches[0].split())
            shien_combined = re.sub(r'\d', '', shien_combined)  
            
            # 年月分の単語の後にある数字最大2桁を抽出（月抽出）
            nengetsu_pattern = re.compile(r'年月分(\d{1,2})')
            nengetsu_matches = nengetsu_pattern.findall(content)
            
            # 枚中の単語の前にある数字最大2桁を抽出（年抽出）
            wareki_pattern = re.compile(r'(\d{1,2})\s*枚中')
            wareki_matches = wareki_pattern.findall(content)
            
            # 数字の単語があった場合で、その次の単語が月火水木金を抽出（日付と曜日抽出）
            pattern1 = re.compile(r'(\d+)\s+([月火水木金])')
            matches1 = pattern1.findall(content)
            
            # 数字と月火水木金が合体している場合も抽出（日付と曜日例外用）
            pattern2 = re.compile(r'(\d+)([月火水木金])')
            matches2 = pattern2.findall(content)

            # :マークの前後にある数字最大2桁の項目を抽出（出退勤の時間抽出）
            colon_pattern = re.compile(r'(\d{1,2}):(\d{1,2})')
            colon_matches = colon_pattern.findall(content)

            # 出退勤の項目数が奇数の場合はエラーメッセージを追加して次に進む
            # 出勤か退勤どちらかが無い場合項目数が奇数の可能性が高いので弾く
            if len(colon_matches) % 2 != 0:
                error_messages.append(f"{shien_combined}さんはpdf内の出退勤情報にエラーがあります")
                continue

            #送迎加算情報取り出し
            esc_pattern = re.compile(r'(\d{1,2}):(\d{1,2})\s+(\d{1,2}):(\d{1,2})(?:\s+(\d+))?')
            esc_matches = esc_pattern.findall(content)

            #抽出した氏名で指定フォルダ内のエクセルファイル検索
            excel_files = [f for f in os.listdir(excel_folder) if os.path.isfile(os.path.join(excel_folder, f)) and shien_combined in f]

            # PDFに氏名があり、エクセルが無い新規の人は新規用テンプレからまずエクセル作成
            if not excel_files:
                # 新規用テンプレートエクセルを選択
                template_filename = os.path.join(excel_folder, '新規用.xlsx')
                
                #該当氏名エクセルファイルが存在しない、新規用.xlsxもフォルダに無い場合はエラーを追加して次へ
                if not os.path.isfile(template_filename):
                    error_messages.append(f"{shien_combined}さんのエクセルが存在しません。新規用テンプレートも存在しません。")
                    continue

                wb = load_workbook(template_filename)

                # 新規の人の名前のエクセルファイルを作成してデータ処理
                filename = '実績_' + shien_combined + '.xlsx'
                filepath = os.path.join(excel_folder, filename)

            #既存ファイルがある継続通所の人はエクセルファイルをロードのみ
            else:
                filename = excel_files[0]
                filepath = os.path.join(excel_folder, filename)
                wb = load_workbook(filepath)

            # コピー元のシートを取得する
            source_sheet = wb['就労継続支援']

            # シート名の月を設定
            new_sheet_name = f"{nengetsu_matches[0]}月"

            # シートのコピーを作成する、シート名は〇月、強制的に追加
            new_sheet = wb.copy_worksheet(source_sheet)
            new_sheet.title = new_sheet_name

            # コピーしたシートを「就労継続支援」の左側に移動する
            wb._sheets.insert(wb._sheets.index(source_sheet), wb._sheets.pop())

            # 年月データを指定セルに書き込む
            new_sheet['D2'] = f"令和{wareki_matches[0]}年"
            new_sheet['L2'] = int(nengetsu_matches[0])

            # 出勤日をC11セルから、曜日をF11セルから、縦に書き込む
            for idx, (day, weekday) in enumerate(matches1 + matches2):
                new_sheet.cell(row=11 + idx, column=3, value=int(day))
                new_sheet.cell(row=11 + idx, column=6, value=weekday)

            # 勤務時間合計を計算するための変数
            total_working_hours = 0

            # BD2セルに値があるかチェック（休憩時間計算の例外チェック）
            has_break_exception = new_sheet['BD2'].value is not None

            # 取得しておいた出勤時間をN11から、退勤時間をS11から、2列にわけて時刻型で書き込む
            for i in range(0, len(colon_matches), 2):
                row_idx = 11 + (i // 2)
                
                # 出勤時間を設定
                cell_n = new_sheet.cell(row=row_idx, column=14)
                start_time = time(int(colon_matches[i][0]), int(colon_matches[i][1]))
                cell_n.value = start_time
                cell_n.number_format = 'HH:MM'
                
                # 退勤時間を設定
                if i + 1 < len(colon_matches):
                    cell_s = new_sheet.cell(row=row_idx, column=19)
                    end_time = time(int(colon_matches[i + 1][0]), int(colon_matches[i + 1][1]))
                    cell_s.value = end_time
                    cell_s.number_format = 'HH:MM'
                    
                    # 勤務時間を計算
                    # 時間を分単位で計算
                    start_minutes = start_time.hour * 60 + start_time.minute
                    end_minutes = end_time.hour * 60 + end_time.minute
                    
                    # 時間差を分単位で計算
                    diff_minutes = end_minutes - start_minutes
                    
                    # 休憩時間の処理
                    # BD2に書き込みがない場合で11:30以前の出勤かつ12:30以降の退勤の場合は1時間差し引く
                    if not has_break_exception:
                        break_start = 11 * 60 + 30  # 11:30 を分単位で表現
                        break_end = 12 * 60 + 30    # 12:30 を分単位で表現
                        
                        if start_minutes <= break_start and end_minutes >= break_end:
                            diff_minutes -= 60  # 1時間(60分)の休憩時間を引く
                    
                    # 分を時間に変換
                    hours_diff = diff_minutes / 60
                    
                    # 15分刻みで切り上げる処理
                    # 例: 5.18時間 → 5.25時間に調整
                    quarter_hours = round(hours_diff * 4) / 4
                    
                    # 勤務時間を合計に加算
                    total_working_hours += quarter_hours
            
            # 合計勤務時間をR41に書き込む
            cell_total = new_sheet.cell(row=41, column=18)
            cell_total.value = total_working_hours
            cell_total.number_format = '0.00'  # 小数点以下2桁で表示
            
            # 書き込みが終わったらエクセルファイルを保存する
            wb.save(filepath)

            # 進捗バーを更新
            progress_bar['value'] += 1
            root.update_idletasks()
            
            # 書き込みが終わったらエクセルファイルを保存する
            wb.save(filepath)

            # 進捗バーを更新
            progress_bar['value'] += 1
            root.update_idletasks()

    #想定外エラー用
    except Exception as e:
        error_messages.append(f"エラーが発生しました: {e}")
    
    #全ての処理終了後PDFを閉じる
    finally:
        pdf_file.close()

    # 処理に失敗した人のエラー一覧表示
    if error_messages:
        for message in error_messages:
            print(message)
        messagebox.showerror("エラー", "\n".join(error_messages))
    else:
        messagebox.showinfo("完了", "処理が完了しました")
        progress_label.config(text="処理が完了しました")

# GUIの設定
root = tk.Tk()
root.title("pdf→エクセル実績表オート記入アプリ")

# 注意書き
notice = tk.Label(root, text="処理前には必ずバックアップを取り、コピーしたエクセルで処理をして下さい。")
notice.pack()
notice = tk.Label(root, text="処理をする際はpdfとエクセルの名前、両方を一致させて下さい。")
notice.pack()
notice = tk.Label(root, text="エクセル側のファイル名はスペースがある場合削除してください。")
notice.pack()
notice = tk.Label(root, text="pdf内の名前とエクセルファイルの名前に不一致やスペースがあったりすると、新規扱いになります")
notice.pack()
notice = tk.Label(root, text="新規の人はファイル名に西暦、ファイル内に名前、受給者番号などを処理完了後入れて下さい。")
notice.pack()
notice = tk.Label(root, text="送迎加算には対応していないので、加算部分だけ手入力でお願いします")
notice.pack()

# PDF選択ボタン、pdf関数呼び出し
pdf_button = tk.Button(root, text="PDFファイルを選択", command=select_pdf)
pdf_button.pack()

# PDFファイルラベル
pdf_label = tk.Label(root, text="選択されたPDF: なし")
pdf_label.pack()

# エクセルフォルダ選択ボタン、エクセル用関数呼び出し
excel_button = tk.Button(root, text="エクセルフォルダを選択", command=select_excel_folder)
excel_button.pack()

# エクセルフォルダラベル
excel_label = tk.Label(root, text="選択されたフォルダ: なし")
excel_label.pack()

# 実行ボタン、実行関数呼び出して処理開始
execute_button = tk.Button(root, text="実行", command=process_files)
execute_button.pack()

# 進捗バー
progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress_bar.pack()

# 進捗ラベル
progress_label = tk.Label(root, text="")
progress_label.pack()

# メインループ
root.mainloop()